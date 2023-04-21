from openpyxl.utils import get_column_letter
import time
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from pynput.keyboard import Key, Controller
import aiohttp
import asyncio
import sys
from aiolimiter import AsyncLimiter
import threading
import configparser
from tkinter import *
from tkinter import filedialog as fd

keyboard = Controller()

if sys.version_info[0] == 3 and sys.version_info[1] >= 8 and sys.platform.startswith('win'):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


def build_tree():
    global treeview
    cols = ("Username", "Password", "IGN", "SERVER", "RANK", "WR", "Games")
    treeview = ttk.Treeview(treeframe, show="headings", yscrollcommand=treescroll.set, columns=cols, height=20)
    treeview.column("Username", width=100)
    treeview.column("Password", width=100)
    treeview.column("IGN", width=100)
    treeview.column("SERVER", width=80)
    treeview.column("RANK", width=100)
    treeview.column("WR", width=50)
    treeview.column("Games", width=80)
    treeview.pack()
    treescroll.config(command=treeview.yview)
    treeview.bind('<ButtonRelease-1>', selectItem)


def load():
    global filepath, col, rowmin, rowmax, api_key, window, usern,passw

    col = 3
    rowmin = 2
    rowmax = 50
    config = configparser.ConfigParser()
    config.read_file(open(r'config.txt'))
    api_key = config.get('Config', 'api_key')
    usern= 'ccd'
    passw = 'aabb'


def run_update_all():
    update_button["state"] = "disabled"
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(update_all_async())
    loop.close()
    treeview.destroy()
    build_tree()
    threading.Thread(target=load_data).start()
    update_button["state"] = "normal"


async def update_all_async():

    RATE_LIMIT_IN_SECOND = 20
    RATE_LIMIT_IN_MINUTES = 100
    rate_limit_sec = AsyncLimiter(RATE_LIMIT_IN_SECOND, 1.0)
    rate_limit_min = AsyncLimiter(RATE_LIMIT_IN_MINUTES, 60.0)
    current_acc_label = ttk.Label(widgets_frame)
    current_acc_label.grid(row=8, column=0, padx=5, pady=5, sticky='nsew')

    print('Updating !')
    async with rate_limit_min:
        async with rate_limit_sec:
            async with aiohttp.ClientSession() as session:

                filepath = 'league.xlsx'
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active

                for row in range(rowmin, rowmax):

                    games = None
                    winratio = None
                    ign = ws[get_column_letter(3) + str(row)].value
                    region = ws[get_column_letter(4) + str(row)].value
                    if ign is None:
                        continue
                    if ign == '--':
                        continue
                    current_acc_label["text"] = ('Updating ' + ign)
                    ign = ign.replace(" ", "")

                    if ws[get_column_letter(4) + str(row)].value == None:
                        continue

                    if region == 'EUW':
                        region = 'euw1'
                    if region == 'EUNE':
                        region = 'eun1'
                    if region == 'RU':
                        region = 'ru'

                    api_url = "https://" + region + ".api.riotgames.com/lol/summoner/v4/summoners/by-name/"

                    api_url = api_url + ign

                    api_url = api_url + '?api_key=' + api_key

                    resp = await session.get(api_url, ssl=False)

                    player_info = await resp.json()
                    if 'id' in player_info:

                        resp = await session.get("https://"+ region +".api.riotgames.com/lol/league/v4/entries/by-summoner/" + player_info['id'] + '?api_key=' + api_key)
                        player_ranked_info = await resp.json()

                        if len(player_ranked_info) != 0:
                            games = int(player_ranked_info[0]['wins']) + int(player_ranked_info[0]['losses'])
                            winratio = str((round((player_ranked_info[0]['wins'] / (
                                    player_ranked_info[0]['wins'] + player_ranked_info[0]['losses'])) * 100))) + "%"
                            rank = player_ranked_info[0]['tier'] + " " + player_ranked_info[0]['rank']

                        else:
                            player_rank = "UNRANKED"

                            rank = player_rank
                            ws[get_column_letter(col + 2) + str(row)] = rank

                        ''' change values in spreadsheet '''
                        ws[get_column_letter(col + 4) + str(row)] = games
                        ws[get_column_letter(col + 3) + str(row)] = winratio
                        ws[get_column_letter(col + 2) + str(row)] = rank
    current_acc_label.destroy()
    wb.save(filepath)
    wb.close()
    print('Done !')


def add_account():

    treeview.destroy()
    username = name_entry.get()
    password = name_entry2.get()
    ign = name_entry3.get()
    region = server_combobox.get()


    path = 'league.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [username, password, ign, region, None, None]
    sheet.append(row_values)
    workbook.save(path)

    name_entry.delete(0,"end")
    name_entry.insert(0,"Username")
    name_entry2.delete(0, "end")
    name_entry2.insert(0, "Password")
    name_entry3.delete(0, "end")
    name_entry3.insert(0, "In Game Name")
    # requested that I disable this:
    # server_combobox.set(combo_list[0])
    build_tree()
    load_data()


def import_from_txt(filename):

    if filename is None:
        return

    path = 'league.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    f = open(filename, 'r')
    text = f.read()
    f.close()
    for line in text.strip().splitlines():
        line = line.split(':')
        username = line[0]
        password = line[1]
        ign = line[2]
        region = line[3]
        row_values = [username, password, ign, region, None, None]
        sheet.append(row_values)
    workbook.save(path)
    treeview.destroy()
    build_tree()
    load_data()


def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


def import_file_func():
    global file_selected
    file_selected = fd.askopenfilename()


def import_messagebox():
    global pop
    pop = Toplevel(root)
    pop.title("Careful !")
    file = None
    pop_label = Label(pop, text="Importing will only work if the accounts are formatted like:\n\n username:password:ign:region")
    pop_label.pack(pady=10)
    pop_frame = Frame(pop)
    pop_frame.pack(pady=5)
    import_file = ttk.Button(pop_frame, text="Select File", command=lambda: [import_file_func()])
    import_file.grid(row=0, column=0, padx=5, pady=5)
    import_button = ttk.Button(pop_frame, text="Import", command=lambda: [import_from_txt(file_selected)])
    import_button.grid(row=1, column=0, padx=5, pady=5)


def selectItem(self):
    global usern, passw
    curItem = treeview.focus()
    usern = treeview.item(curItem)['values'][0]
    passw = treeview.item(curItem)['values'][1]


def login(username, password):


    if username != "ccd":
        #TODO new login method using api
        config = configparser.ConfigParser()
        config.read_file(open(r'config.txt'))
        path = config.get('Config', 'path')
        subprocess.call([path])
        time.sleep(4)

        keyboard.type(username)
        keyboard.press(Key.tab)
        keyboard.release(Key.tab)
        time.sleep(0.25)
        keyboard.type(password)

        keyboard.press(Key.enter)
        keyboard.release(Key.enter)

    else:
        messagebox.showinfo(title='Login Error', message='Choose an account first !')


def load_data():
    path = 'league.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text = col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)
    workbook.close()


def delete_acc():
    target_row = None
    filepath = 'league.xlsx'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    curItem = treeview.focus()
    usern = treeview.item(curItem)['values'][0]

    '''FINDS THE ROW OF THE SELECTED IGN'''
    for row in range(rowmin, rowmax):
        userncol = get_column_letter(col-2)
        ign = ws[userncol + str(row)].value

        if ign == usern:
            target_row = row
            break
    ws.delete_rows(idx=target_row)
    wb.save(filepath)
    treeview.destroy()
    build_tree()
    load_data()


def delete_if_empty(field):
    if field.get() == "Username":
        field.delete('0', 'end')
    if field.get() == "Password":
        field.delete('0', 'end')
    if field.get() == "In Game Name":
        field.delete('0', 'end')
    else:
        return



def gui():
        global cols, treeframe, treescroll, update_button, name_entry, name_entry2, name_entry3, server_combobox, combo_list, mode_switch, style, widgets_frame, root
        root = tk.Tk()
        root.title("league pwm")
        style = ttk.Style(root)
        root.tk.call("source", r"forest-light.tcl")
        root.tk.call("source", r"forest-dark.tcl")
        style.theme_use("forest-dark")

        combo_list = ["EUW", "EUNE", "RU"]

        frame = ttk.Frame(root)
        frame.pack()
        widgets_frame = ttk.LabelFrame(frame, text="Add Account")
        widgets_frame.grid(row=0, column=0, padx=20, pady=10)

        name_entry = ttk.Entry(widgets_frame)
        name_entry.insert(0, "Username")
        name_entry.bind("<FocusIn>", lambda e: delete_if_empty(name_entry))
        name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        name_entry2 = ttk.Entry(widgets_frame)
        name_entry2.insert(0, "Password")
        name_entry2.bind("<FocusIn>", lambda e: delete_if_empty(name_entry2))
        name_entry2.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        name_entry3 = ttk.Entry(widgets_frame)
        name_entry3.insert(0, "In Game Name")
        name_entry3.bind("<FocusIn>", lambda e: delete_if_empty(name_entry3))
        name_entry3.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        server_combobox = ttk.Combobox(widgets_frame, values=combo_list)
        server_combobox.current(0)
        server_combobox.grid(row=3, column=0, padx=5, pady=5, sticky='ew')

        add_button = ttk.Button(widgets_frame, text="Add", command=add_account)
        add_button.grid(row=4, column=0, padx=5, pady=5, sticky='nsew')

        separator = ttk.Separator(widgets_frame)
        separator.grid(row=6, column=0, padx=(20, 10), pady=10, sticky='ew')

        treeframe = ttk.Frame(frame)
        treeframe.grid(row=0, column=1, pady=10)
        treescroll = ttk.Scrollbar(treeframe)

        #Grid.rowconfigure(treeframe, 0, weight=1)
        #Grid.columnconfigure(treeframe, 0, weight=1)

        treescroll.pack(side="right", fill="y")

        cols = ("Username", "Password", "IGN", "SERVER", "RANK", "WR", "Games")

        del_button = ttk.Button(widgets_frame, text="Delete Selected", command=lambda: [delete_acc()])
        del_button.grid(row=5, column=0, padx=5, pady=5, sticky='nsew')

        update_button = ttk.Button(widgets_frame, text="Update", command=lambda: [threading.Thread(target=run_update_all).start()])
        update_button.grid(row=7, column=0, padx=5, pady=5, sticky='nsew')

        login_button = ttk.Button(widgets_frame, text="Login", command=lambda: login(usern, passw))
        login_button.grid(row=9, column=0, padx=5, pady=5, sticky='nsew')

        bottom_frame = ttk.Frame(widgets_frame)
        bottom_frame.grid(row=10, column=0, pady=10)

        mode_switch = ttk.Checkbutton(bottom_frame, text="Mode", style="Switch", command=toggle_mode)
        mode_switch.grid(row=0, column=0, padx=5, pady=10, sticky="nsew")

        import_txt = ttk.Button(bottom_frame, text="Import", command=import_messagebox)
        import_txt.grid(row=0, column=1, padx=5, pady=5, sticky='nsew')

        load()
        build_tree()
        load_data()

        root.mainloop()


gui()
