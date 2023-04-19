from openpyxl.utils import get_column_letter
import time
import subprocess
import tkinter as tk
from tkinter import ttk
import openpyxl
from pynput.keyboard import Key, Controller
import aiohttp
import asyncio
import sys
from aiolimiter import AsyncLimiter

keyboard = Controller()

if sys.version_info[0] == 3 and sys.version_info[1] >= 8 and sys.platform.startswith('win'):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

def build_tree():
    global treeview
    cols = ("Username", "Password", "IGN", "SERVER", "RANK", "WR", "Games")
    treeview = ttk.Treeview(treeframe, show="headings", yscrollcommand=treescroll.set, columns=cols, height=20)
    treeview.column("Username", width=100)
    treeview.column("Password", width=100)
    treeview.column("IGN", width=100)
    treeview.column("SERVER", width=80)
    treeview.column("RANK", width=100)
    treeview.column("WR", width=50)
    treeview.column("Games", width=80)
    treeview.pack()
    treescroll.config(command=treeview.yview)
    treeview.bind('<ButtonRelease-1>', selectItem)


def delete_tree():
    treeview.destroy()

def reg_string(region):
    if region == 'EUW':
        region = 'euw1'
    elif region == 'EUNE':
        region ='eun1'
    elif region == 'RU':
        region ='ru'


def load():
    global filepath, col, rowmin, rowmax, api_key, window, usern,passw

    col = 3
    rowmin = 2
    rowmax = 50
    f = open('api.txt', 'r')
    api_key = f.read()
    f.close()
    usern= 'ccd'
    passw = 'aabb'


def run_update_all():
    loop = asyncio.get_event_loop()
    loop.run_until_complete(update_all_async())
    loop.close()


async def update_all_async():

    RATE_LIMIT_IN_SECOND = 20
    RATE_LIMIT_IN_MINUTES = 100
    rate_limit_sec = AsyncLimiter(RATE_LIMIT_IN_SECOND, 1.0)
    rate_limit_min = AsyncLimiter(RATE_LIMIT_IN_MINUTES, 60.0)
    update_button["state"] = "disabled"
    print('Updating !')
    async with rate_limit_min:
        async with rate_limit_sec:
            async with aiohttp.ClientSession() as session:

                filepath = 'league.xlsx'
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active

                f = open('api.txt', 'r')
                api_key = f.read()
                f.close()

                for row in range(rowmin, rowmax):

                    games = None
                    winratio = None
                    ign = ws[get_column_letter(3) + str(row)].value
                    region = ws[get_column_letter(4) + str(row)].value
                    if ign is None:
                        continue
                    if ign == '--':
                        continue
                    print('Updating ' + ign)
                    ign = ign.replace(" ", "")

                    if ws[get_column_letter(4) + str(row)].value == None:
                        continue

                    if region == 'EUW':
                        region = 'euw1'
                    if region == 'EUNE':
                        region = 'eun1'
                    if region == 'RU':
                        region = 'ru'

                    api_url = "https://" + region + ".api.riotgames.com/lol/summoner/v4/summoners/by-name/"

                    api_url = api_url + ign

                    api_url = api_url + '?api_key=' + api_key

                    resp = await session.get(api_url, ssl=False)

                    player_info = await resp.json()
                    if 'id' in player_info:

                        resp = await session.get("https://"+ region +".api.riotgames.com/lol/league/v4/entries/by-summoner/" + player_info['id'] + '?api_key=' + api_key)
                        player_ranked_info = await resp.json()

                        if len(player_ranked_info) != 0:
                            games = int(player_ranked_info[0]['wins']) + int(player_ranked_info[0]['losses'])
                            winratio = str((round((player_ranked_info[0]['wins'] / (
                                    player_ranked_info[0]['wins'] + player_ranked_info[0]['losses'])) * 100))) + "%"
                            rank = player_ranked_info[0]['tier'] + " " + player_ranked_info[0]['rank']

                        else:
                            player_rank = "UNRANKED"

                            rank = player_rank
                            ws[get_column_letter(col + 2) + str(row)] = rank

                        ''' change values in spreadsheet '''
                        ws[get_column_letter(col + 4) + str(row)] = games
                        ws[get_column_letter(col + 3) + str(row)] = winratio
                        ws[get_column_letter(col + 2) + str(row)] = rank

    wb.save(filepath)
    wb.close()
    delete_tree()
    build_tree()
    load_data()
    update_button["state"] = "normal"
    print('Done !')


def add_account():

    delete_tree()
    username = name_entry.get()
    password = name_entry2.get()
    ign = name_entry3.get()
    region = server_combobox.get()


    path = 'league.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [username, password, ign, region, None, None]
    sheet.append(row_values)
    workbook.save(path)

    name_entry.delete(0,"end")
    name_entry.insert(0,"Username")
    name_entry2.delete(0, "end")
    name_entry2.insert(0, "Password")
    name_entry3.delete(0, "end")
    name_entry3.insert(0, "In Game Name")
    server_combobox.set(combo_list[0])

    build_tree()
    load_data()


def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


def selectItem(self):
    global usern, passw
    curItem = treeview.focus()
    usern = treeview.item(curItem)['values'][0]
    passw = treeview.item(curItem)['values'][1]


def login(username, password):

    f = open('leaguepath.txt', 'r')
    path = f.read()
    subprocess.call([path,  '--launch-patchline=live', '--allow-multiple-clients'])
    f.close()
    time.sleep(4)

    keyboard.type(username)
    keyboard.press(Key.tab)
    keyboard.release(Key.tab)
    time.sleep(0.25)
    keyboard.type(password)

    keyboard.press(Key.enter)
    keyboard.release(Key.enter)


def load_data():
    path = 'league.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text = col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)
    workbook.close()


def delete_acc():

    filepath = 'league.xlsx'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    curItem = treeview.focus()
    usern = treeview.item(curItem)['values'][0]

    '''FINDS THE ROW OF THE SELECTED IGN'''
    for row in range(rowmin, rowmax):
        userncol = get_column_letter(col-2)
        ign = ws[userncol + str(row)].value

        if ign == usern:
            target_row = row
            break
    ws.delete_rows(idx=target_row)
    wb.save(filepath)
    delete_tree()
    build_tree()
    load_data()



def gui():
        global cols, treeframe, treescroll, update_button, name_entry, name_entry2, name_entry3, server_combobox, combo_list, mode_switch, style
        root = tk.Tk()
        root.title("league pwm")
        style = ttk.Style(root)
        root.tk.call("source", r"forest-light.tcl")
        root.tk.call("source", r"forest-dark.tcl")
        style.theme_use("forest-dark")

        combo_list = ["EUW", "EUNE", "RU"]

        frame = ttk.Frame(root)
        frame.pack()

        widgets_frame = ttk.LabelFrame(frame, text="Add Account")
        widgets_frame.grid(row=0, column=0, padx=20, pady=10)

        name_entry = ttk.Entry(widgets_frame)
        name_entry.insert(0, "Username")
        name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
        name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        name_entry2 = ttk.Entry(widgets_frame)
        name_entry2.insert(0, "Password")
        name_entry2.bind("<FocusIn>", lambda e: name_entry2.delete('0', 'end'))
        name_entry2.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        name_entry3 = ttk.Entry(widgets_frame)
        name_entry3.insert(0, "In Game Name")
        name_entry3.bind("<FocusIn>", lambda e: name_entry3.delete('0', 'end'))
        name_entry3.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        server_combobox = ttk.Combobox(widgets_frame, values=combo_list)
        server_combobox.current(0)
        server_combobox.grid(row=3, column=0, padx=5, pady=5, sticky='ew')

        add_button = ttk.Button(widgets_frame, text="Add", command=add_account)
        add_button.grid(row=4, column=0, padx=5, pady=5, sticky='nsew')

        separator = ttk.Separator(widgets_frame)
        separator.grid(row=6, column=0, padx=(20, 10), pady=10, sticky='ew')

        treeframe = ttk.Frame(frame)
        treeframe.grid(row=0, column=1, pady=10)
        treescroll = ttk.Scrollbar(treeframe)
        treescroll.pack(side="right", fill="y")

        cols = ("Username", "Password", "IGN", "SERVER", "RANK", "WR", "Games")



        del_button = ttk.Button(widgets_frame, text="Delete Selected", command=lambda: [delete_acc()])
        del_button.grid(row=5, column=0, padx=5, pady=5, sticky='nsew')

        update_button = ttk.Button(widgets_frame, text="Update", command=lambda: [run_update_all()])
        update_button.grid(row=7, column=0, padx=5, pady=5, sticky='nsew')

        login_button = ttk.Button(widgets_frame, text="Login", command=lambda: login(usern, passw))
        login_button.grid(row=8, column=0, padx=5, pady=5, sticky='nsew')

        mode_switch = ttk.Checkbutton(widgets_frame, text="Mode", style="Switch", command=toggle_mode)
        mode_switch.grid(row=9, column=0, padx=5, pady=10, sticky="nsew")

        load()
        build_tree()
        load_data()

        root.mainloop()



gui()


